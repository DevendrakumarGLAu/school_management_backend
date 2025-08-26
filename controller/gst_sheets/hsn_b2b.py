from openpyxl.styles import PatternFill, Font, Alignment

def generate_hsn_b2b_sheet(wb):
    ws = wb.create_sheet(title="hsn(b2b)")

    # Styles
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # === Row 1: Title and HELP ===
    ws['A1'] = "Summary For HSN(12)"
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center
    ws.merge_cells('A1:J1')
    ws['K1'] = "HELP"
    ws['K1'].font = font_bold
    ws['K1'].alignment = align_center

    # === Row 2: Summary Headers ===
    headers_row2 = ["No. of HSN", "", "", "", "Total Value", "", "Total Taxable Value", 
                    "Total Integrated Tax", "Total Central Tax", "Total State/UT Tax", "Total Cess"]

    for i, val in enumerate(headers_row2, start=1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.fill = blue_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # === Row 3: Summary values (all zero for now) ===
    ws.append([0, "", "", "", 0.00, "", 0.00, 0.00, 0.00, 0.00, 0.00])

    # === Row 4: Table Headers ===
    headers_row4 = [
        "HSN", "Description", "UQC", "Total Quantity", "Total Value", "Rate", "Taxable Value",
        "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"
    ]

    for i, val in enumerate(headers_row4, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    # Optional: Set column widths for better readability
    column_widths = [10, 25, 12, 18, 15, 10, 18, 20, 20, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = width
