from openpyxl.styles import PatternFill, Font, Alignment

def generate_b2cl_sheet(wb):
    ws = wb.create_sheet(title="b2cl")

    # === Styles ===
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # === Row 1: Title + HELP ===
    ws['A1'] = "Summary For B2CL(5)"
    ws.merge_cells('A1:H1')
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center

    ws['I1'] = "HELP"
    ws['I1'].font = font_bold
    ws['I1'].alignment = align_center

    # === Row 2: Summary Header ===
    ws['A2'] = "No. of Invoices"
    ws['A2'].fill = blue_fill
    ws['A2'].font = font_white_bold
    ws['A2'].alignment = align_center

    # === Row 3: Summary Values ===
    ws.append([0, "", 0.00, "", "", "", 0.00, "", ""])

    # === Row 4: Table Headers ===
    headers = [
        "Invoice Number", "Invoice date", "Invoice Value", "Place Of Supply",
        "Applicable % of Tax Rate", "Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"
    ]
    for i, val in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    # Optional: Set column widths
    widths = [20, 15, 18, 22, 25, 10, 18, 15, 25]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
