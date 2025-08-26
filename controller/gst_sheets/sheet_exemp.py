from openpyxl.styles import PatternFill, Font, Alignment

def generate_exemp_sheet(wb):
    ws = wb.create_sheet(title="exemp")

    # ========== Styles ==========
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # ========== Row 1: Title ==========
    ws['A1'] = "Summary For Nil rated, exempted and non GST outward supplies (8)"
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center
    ws.merge_cells('A1:C1')

    ws['D1'] = "HELP"
    ws['D1'].font = font_bold
    ws['D1'].alignment = align_center

    # ========== Row 2: Header Totals ==========
    ws['B2'] = "Total Nil Rated Supplies"
    ws['C2'] = "Total Exempted Supplies"
    ws['D2'] = "Total Non-GST Supplies"

    for col in ['B', 'C', 'D']:
        cell = ws[f'{col}2']
        cell.fill = blue_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # ========== Row 3: Totals Values ==========
    ws['B3'] = 0.00
    ws['C3'] = 0.00
    ws['D3'] = 0.00

    # ========== Row 4: Column Headers for Table ==========
    ws['A4'] = "Description"
    ws['B4'] = "Nil Rated Supplies"
    ws['C4'] = "Exempted(other than nil rated/non GST supply)"
    ws['D4'] = "Non-GST Supplies"

    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}4']
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    # ========== Row 5-8: Table Data ==========
    table_data = [
        ["Inter-State supplies to registered persons", 0.00, "", ""],
        ["Intra-State supplies to registered persons", 0.00, "", ""],
        ["Inter-State supplies to unregistered persons", 7080.00, "", ""],
        ["Intra-State supplies to unregistered persons", 4269.00, "", ""],
    ]

    for row in table_data:
        ws.append(row)

    # Optional: Adjust column widths for better readability
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
