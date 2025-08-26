from openpyxl.styles import PatternFill, Font, Alignment

def generate_sez_sheet(wb):
    ws = wb.create_sheet(title="b2b,sez,de")

    # === Styles ===
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # === Row 1: Title + HELP ===
    ws['A1'] = "Summary For B2B, SEZ, DE (4A, 4B, 6B, 6C)"
    ws.merge_cells('A1:L1')
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center

    ws['M1'] = "HELP"
    ws['M1'].font = font_bold
    ws['M1'].alignment = align_center

    # === Row 2: Summary Headers ===
    headers_row2 = [
        "No. of Recipients", "", "No. of Invoices", "", "Total Invoice Value", "", "", "", "", "", "",
        "Total Taxable Value", "Total Cess"
    ]
    for i, val in enumerate(headers_row2, start=1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.fill = blue_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # === Row 3: Summary Values (Zeros) ===
    ws.append([0, "", 0, "", 0.00, "", "", "", "", "", "", 0.00, 0.00])

    # === Row 4: Table Headers ===
    headers_row4 = [
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice date", "Invoice Value",
        "Place Of Supply", "Reverse Charge", "Applicable % of Tax Rate", "Invoice Type",
        "E-Commerce GSTIN", "Rate", "Taxable Value", "Cess Amount"
    ]
    for i, val in enumerate(headers_row4, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    # === Column Widths (Optional) ===
    widths = [25, 22, 20, 15, 18, 20, 15, 25, 15, 25, 10, 18, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
