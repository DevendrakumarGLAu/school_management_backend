from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

def generate_b2cs_sheet(wb, tcs_sales_return_df, tcs_sales_df, tax_invoice_details_df, gstNumber):
    ws = wb.create_sheet(title="b2cs")

    # === Styles ===
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # === Row 1: Title and HELP ===
    ws['A1'] = "Summary For B2CS(7)"
    ws.merge_cells('A1:F1')
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center

    ws['G1'] = "HELP"
    ws['G1'].font = font_bold
    ws['G1'].alignment = align_center

    # === Row 2: Summary Headers ===
    ws['D2'] = "Total Taxable  Value"
    ws['F2'] = "Total Cess"
    ws['D2'].fill = blue_fill
    ws['D2'].font = font_white_bold
    ws['D2'].alignment = align_center
    ws['F2'].fill = blue_fill
    ws['F2'].font = font_white_bold
    ws['F2'].alignment = align_center

    # === GST State Code Map ===
    gst_state_codes = {
        "jammu & kashmir": "01", 
        "himachal pradesh": "02",
        "punjab": "03",
        "chandigarh": "04",
        "uttarakhand": "05", 
        "haryana": "06",
        "delhi": "07",
        "rajasthan": "08", 
        "uttar pradesh": "09",
        "bihar": "10",
        "sikkim": "11", 
        "arunachal pradesh": "12", 
        "nagaland": "13", "manipur": "14",
        "mizoram": "15", "tripura": "16", "meghalaya": "17", "assam": "18", "west bengal": "19",
        "jharkhand": "20", "odisha": "21", "chhattisgarh": "22", "madhya pradesh": "23", "gujarat": "24",
        "daman and diu": "25", "dadra and nagar haveli and daman and diu": "26", "maharashtra": "27",
        "andhra pradesh (old)": "28", "karnataka": "29", "goa": "30", "lakshadweep": "31", "kerala": "32",
        "tamil nadu": "33", "pondicherry": "34", "andaman and nicobar islands": "35", "telangana": "36",
        "andhra pradesh": "37", "ladakh": "38","puducherry": "34",
    }

    def get_state_code(state_name):
        if not isinstance(state_name, str):
            return None
        return gst_state_codes.get(state_name.strip().lower())

    # === Group sales and returns ===
    sales_grouped = tcs_sales_df.groupby(['gst_rate', 'end_customer_state_new'])['total_taxable_sale_value'].sum()
    returns_grouped = tcs_sales_return_df.groupby(['gst_rate', 'end_customer_state_new'])['total_taxable_sale_value'].sum()

    # === Create a dictionary for return lookups ===
    returns_dict = returns_grouped.to_dict()

    # === Prepare data rows ===
    data_rows = []
    total_net_taxable = 0

    for (gst_rate, state_name), sales_value in sales_grouped.items():
        return_value = returns_dict.get((gst_rate, state_name), 0)
        net_taxable = sales_value - return_value
        if abs(net_taxable) < 0.01:
            continue
        total_net_taxable += net_taxable

        state_code = get_state_code(state_name)
        if not state_code:
            continue  # Skip if state code not found

        state_display = f"{state_code}-{state_name.title()}"
        data_rows.append((state_display, gst_rate, round(net_taxable, 2)))

    # === Write summary ===
    ws['D3'] = round(total_net_taxable, 2)
    ws['F3'] = 0  # Assuming no cess

    # === Headers ===
    headers = [
        "Type", "Place Of Supply", "Applicable % of Tax Rate", "Rate",
        "Taxable Value", "Cess Amount", "E-Commerce GSTIN"
    ]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    # === Write Data ===
    data_rows.sort(key=lambda x: x[0].split('-', 1)[1].lower())

    row_start = 5
    for i, (state_display, gst_rate, net_taxable) in enumerate(data_rows):
        row = row_start + i
        # print("OE",state_display,gst_rate,net_taxable)
        if state_display== '34-Puducherry':
            print("OE",state_display,gst_rate,net_taxable)
        ws.cell(row=row, column=1, value="OE")
        ws.cell(row=row, column=2, value=state_display)
        ws.cell(row=row, column=3, value="")  # Applicable % of Tax Rate (optional)
        ws.cell(row=row, column=4, value=gst_rate)
        ws.cell(row=row, column=5, value=net_taxable)
        ws.cell(row=row, column=6, value=0)  # Cess assumed zero
        ws.cell(row=row, column=7, value="")  # E-commerce GSTIN

    # === Adjust Column Widths ===
    widths = [12, 22, 25, 10, 18, 15, 25]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return wb
