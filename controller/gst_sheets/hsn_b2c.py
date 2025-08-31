from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def generate_hsn_sheet(wb, tcs_sales_return_df, tcs_sales_df, tax_invoice_details_df,gstNumber):
    gst_state_codes = {
    "jammu & kashmir": "01", "himachal pradesh": "02", "punjab": "03",
    "chandigarh": "04", "uttarakhand": "05", "haryana": "06",
    "delhi": "07", "rajasthan": "08", "uttar pradesh": "09",
    "bihar": "10", "sikkim": "11", "arunachal pradesh": "12",
    "nagaland": "13", "manipur": "14", "mizoram": "15", "tripura": "16",
    "meghalaya": "17", "assam": "18", "west bengal": "19",
    "jharkhand": "20", "odisha": "21", "chhattisgarh": "22",
    "madhya pradesh": "23", "gujarat": "24", "daman and diu": "25",
    "dadra and nagar haveli and daman and diu": "26", "maharashtra": "27",
    "andhra pradesh (old)": "28", "karnataka": "29", "goa": "30",
    "lakshadweep": "31", "kerala": "32", "tamil nadu": "33",
    "pondicherry": "34", "puducherry": "34",
    "andaman and nicobar islands": "35", "telangana": "36",
    "andhra pradesh": "37", "ladakh": "38"
}

    def get_state_code(state_name):
        if not isinstance(state_name, str):
            return None
        return gst_state_codes.get(state_name.strip().lower())
    
    my_state_code = gstNumber[:2]
    
    ws = wb.create_sheet(title="hsn(b2c)")

    # ========== STYLING ==========
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # ========== CLEAN AND PREPARE DATA ==========
    cols_to_clean = ['quantity', 'total_taxable_sale_value', 'total_invoice_value', 'gst_rate']
    for col in cols_to_clean:
        if col in tcs_sales_df.columns:
            tcs_sales_df[col] = pd.to_numeric(
                tcs_sales_df[col].astype(str).str.replace("'", "", regex=False),
                errors='coerce'
            ).fillna(0)
        
        if col in tcs_sales_return_df.columns:
            tcs_sales_return_df[col] = pd.to_numeric(
                tcs_sales_return_df[col].astype(str).str.replace("'", "", regex=False),
                errors='coerce'
            ).fillna(0)
            
    if "end_customer_state" in tcs_sales_return_df.columns:
        tcs_sales_return_df["gst_state_code"] = tcs_sales_return_df["end_customer_state"].apply(get_state_code)
            
    if "end_customer_state" in tcs_sales_df.columns:
        tcs_sales_df["gst_state_code"] = tcs_sales_df["end_customer_state"].apply(get_state_code)
            
    # ========== GROUP SALES AND RETURNS SEPARATELY ==========
    sales_grouped = tcs_sales_df.groupby('hsn_code').agg({
        'quantity': 'sum',
        'total_taxable_sale_value': 'sum',
        'total_invoice_value': 'sum',
        'gst_rate': 'first'
    }).reset_index()

    returns_grouped = tcs_sales_return_df.groupby('hsn_code').agg({
        'quantity': 'sum',
        'total_taxable_sale_value': 'sum',
        'total_invoice_value': 'sum'
    }).reset_index()

    # ========== MERGE SALES AND RETURNS TO COMPUTE NETS ==========
    hsn_grouped = pd.merge(
        sales_grouped,
        returns_grouped,
        on='hsn_code',
        how='left',
        suffixes=('', '_return')
    ).fillna(0)

    # Compute net values
    hsn_grouped['net_quantity'] = hsn_grouped['quantity'] - hsn_grouped['quantity_return']
    hsn_grouped['net_taxable'] = hsn_grouped['total_taxable_sale_value'] - hsn_grouped['total_taxable_sale_value_return']
    hsn_grouped['net_invoice'] = hsn_grouped['total_invoice_value'] - hsn_grouped['total_invoice_value_return']
    
    # Round
    hsn_grouped['net_taxable'] = hsn_grouped['net_taxable'].round(2)
    hsn_grouped['net_invoice'] = hsn_grouped['net_invoice'].round(2)

    # ========== EXCEL SHEET SETUP ==========

    # Row 1 - Title and Help
    ws['A1'] = "Summary For HSN(12)"
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center
    # ws.merge_cells('A1:G1')

    ws['H1'] = "HELP"
    ws['H1'].font = font_bold
    ws['H1'].alignment = align_center

    # Row 2 - Headers for summary
    headers_row2 = ["No. of HSN", "","","", "Total Value", "Total Taxable Value",
                    "Total Integrated Tax", "Total Central Tax", "Total State/UT Tax", "Total Cess"]

    for i, val in enumerate(headers_row2, start=1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.fill = blue_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # ✅ Row 3 - Summary values (calculated dynamically)
    total_value = hsn_grouped['net_invoice'].sum()
    taxable_value = hsn_grouped['net_taxable'].sum()
    # total_integrated_tax = hsn_grouped['Integrated Tax Amount'].sum()
    # total_central_tax = hsn_grouped['Central Tax Amount'].sum()
    # total_state_ut_tax = hsn_grouped['State/UT Tax Amount'].sum()
    # total_cess = hsn_grouped['Cess Amount'].sum()

    # Append summary row
    ws.append([
        hsn_grouped.shape[0], "", "", "",
        total_value, taxable_value,
        0, 0, 0, 0
    ])
    # Row 4 - HSN Detail Headers
    headers_row4 = [
        "HSN", "Description", "UQC", "Total Quantity", "Total Value", "Rate",
        "Taxable Value", "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"
    ]

    for i, val in enumerate(headers_row4, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    for _, row in hsn_grouped.iterrows():
        taxable = row['net_taxable']
        rate = row['gst_rate']
        igst_amt, cgst_amt, sgst_amt = 0.0, 0.0, 0.0
        
        cust_state = get_state_code(row.get("end_customer_state", ""))

        if cust_state and cust_state != my_state_code:  # INTER
           igst_amt = round(taxable * rate / 100, 2)
           cgst_amt=0
           sgst_amt=0
        else:  # INTRA
            igst_amt = 0
            cgst_amt = round(taxable * (rate / 2) / 100, 2)
            sgst_amt = round(taxable * (rate / 2) / 100, 2)

        ws.append([
            int(row['hsn_code']), "", "PCS-PIECES",
            int(row['net_quantity']),
            row['net_invoice'],
            row['gst_rate'],
            row['net_taxable'],
            igst_amt,
            cgst_amt,
            sgst_amt,
            0.00
        ])