from openpyxl.styles import PatternFill, Font, Alignment

def generate_eco_sheet(wb,meesho_gst,supplier_name,Net_value_of_supplies):
    ws = wb.create_sheet(title="eco")

    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    peach_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # ws.merge_cells('A1:H1')
    ws['A1'] = "Summary For Supplies through ECO-14"
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center

    ws['H1'] = "HELP"
    ws['H1'].font = font_bold
    ws['H1'].alignment = align_center

    headers_row2 = ["", "No. of E-Commerce Operator", "", "Total Net Value of Supplies", 
                    "Total Integrated Tax", "Total Central Tax", "Total State/UT Tax", "Total Cess"]

    for i, val in enumerate(headers_row2, start=1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.fill = blue_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # ws['I2'].fill = blue_fill
    # ws['I2'].font = font_white_bold
    # ws['I2'].alignment = align_center

    ws.append(["", 0, "", 0.00, 0.00, 0.00, 0.00, 0.00])

    headers_row4 = ["Nature of Supply", "GSTIN of E-Commerce Operator", "E-Commerce Operator Name",
                    "Net value of supplies", "Integrated tax", "Central tax", "State/UT tax", "Cess"]

    for i, val in enumerate(headers_row4, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = peach_fill
        cell.font = font_bold
        cell.alignment = align_center

    ws.append([
        "Liable to collect tax u/s 52(TCS)", meesho_gst, supplier_name, 
        Net_value_of_supplies, 0.00, 0.00, 0.00, 0.00
    ])
