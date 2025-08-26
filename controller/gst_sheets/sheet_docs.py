from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def generate_docs_sheet(wb, tax_invoice_details_df):
    ws = wb.active
    ws.title = "docs"

    # Styles
    header_fill = PatternFill(start_color='986801', end_color='986801', fill_type='solid')
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # Title Row (Row 1)
    ws['A1'] = "Summary of documents issued during the tax period (13)"
    ws.merge_cells('A1:E1')
    ws['A1'].fill = blue_fill
    ws['A1'].font = font_white_bold
    ws['A1'].alignment = align_center

    # ===== Row 2 & 3: Totals =====
    ws['D2'] = "Total Number"
    ws['E2'] = "Total Cancelled"
    
    total_number = len(tax_invoice_details_df)
    
    # Cancelled logic – assuming there's a "Cancelled" column (adjust if not present)
    if 'Cancelled' in tax_invoice_details_df.columns:
        total_cancelled = tax_invoice_details_df['Cancelled'].astype(bool).sum()
    else:
        total_cancelled = 0  # fallback if no "Cancelled" column
    
    ws['D3'] = total_number
    ws['E3'] = total_cancelled

    # ===== Row 4: Headers =====
    headers = ["Nature of Document", "Sr. No. From", "Sr. No. To", "Total Number", "Cancelled"]
    for i, val in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.fill = header_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # ===== Helper: Natural sort for alphanumeric invoice numbers =====
    def sort_key(val):
        import re
        match = re.match(r"([a-zA-Z]+)(\d+)", str(val))
        if match:
            prefix, num = match.groups()
            return (prefix, int(num))
        return (val, 0)

    # ===== Document Types to Process =====
    document_types = {
        "INVOICE": "Invoices for outward supply",
        "CREDIT NOTE": "Credit Note"
    }

    # ===== Row 5 onward: Document Rows =====
    for doc_type, label in document_types.items():
        filtered_df = tax_invoice_details_df[tax_invoice_details_df['Type'] == doc_type]
        invoice_numbers = sorted(filtered_df['Invoice No.'].dropna().unique(), key=sort_key)

        if invoice_numbers:
            sr_from = invoice_numbers[0]
            sr_to = invoice_numbers[-1]
            total_count = len(invoice_numbers)

            if 'Cancelled' in filtered_df.columns:
                cancelled_count = filtered_df['Cancelled'].astype(bool).sum()
            else:
                cancelled_count = 0

            ws.append([label, sr_from, sr_to, total_count, cancelled_count])
